import frappe
from frappe.utils import getdate, cint

from attendance_processor.utils.processor import (
    get_attendance_records,
    get_missed_requests_lookup,
    get_leave_applications_lookup,
    get_short_leave_lookup,
    get_two_late_lookup,
    analyse_employee,
)
from attendance_processor.utils.email_report import send_summary_email


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _is_employee_active(att_records):
    for rec in att_records:
        if rec.status in ("Present", "Half Day"):
            return True
        if rec.custom_ucsc_status in ("Present", "Half Day"):
            return True
    return False


def _get_active_employee_ids(employee=None):
    """Return a set of employee IDs whose HR status is 'Active'."""
    filters = {"status": "Active"}
    if employee:
        filters["name"] = employee
    return set(frappe.get_all("Employee", filters=filters, pluck="name"))


def _fmt_time(val):
    """Extract HH:MM:SS portion from a datetime/string; return None if absent."""
    if not val:
        return None
    s = str(val).strip()
    return s.split(" ", 1)[1] if " " in s else s


def _serialize_record(rec):
    """Convert a frappe._dict attendance record to a plain JSON-safe dict."""
    return {
        "name":              rec.name,
        "attendance_date":   str(rec.attendance_date) if rec.attendance_date else None,
        "status":            rec.status or "",
        "custom_ucsc_status": rec.custom_ucsc_status or "",
        "in_time":           _fmt_time(rec.in_time),
        "out_time":          _fmt_time(rec.out_time),
        "shift":             rec.shift or "",
        "custom_remarks":    (rec.custom_remarks or "")[:80],
        "employee":          rec.employee,
        "employee_name":     rec.employee_name,
    }


def _build_emp_data(all_records):
    """Group a flat list of attendance records by employee."""
    emp_data = {}
    for rec in all_records:
        if rec.employee not in emp_data:
            emp_data[rec.employee] = {"name": rec.employee_name, "records": []}
        emp_data[rec.employee]["records"].append(rec)
    return emp_data


# ---------------------------------------------------------------------------
# Internal analysis helper (shared by the API endpoints below)
# ---------------------------------------------------------------------------

def _run_analysis(from_date, to_date, employees=None):
    """
    Core attendance analysis.  Returns (results, from_date, to_date) where
    *results* is the same list-of-dicts used by all public endpoints and
    *from_date* / *to_date* are normalised date objects.
    """
    import json
    from_date = getdate(from_date)
    to_date   = getdate(to_date)

    employees_list = None
    if employees:
        raw = json.loads(employees) if isinstance(employees, str) else employees
        employees_list = [e for e in raw if e] or None

    missed_lookup      = get_missed_requests_lookup(from_date, to_date)
    leave_lookup       = get_leave_applications_lookup(from_date, to_date)
    short_leave_lookup = get_short_leave_lookup(from_date, to_date)
    two_late_lookup    = get_two_late_lookup(from_date, to_date)
    all_records        = get_attendance_records(from_date, to_date,
                                                employees=employees_list)

    emp_data = _build_emp_data(all_records)

    # Enforce ERPNext User Permissions: frappe.get_list applies role permissions
    # and User Permissions automatically, so non-privileged users only see the
    # Employee records they are permitted to access.
    if "System Manager" not in frappe.get_roles():
        permitted_ids = set(frappe.get_list("Employee", pluck="name"))
        emp_data = {k: v for k, v in emp_data.items() if k in permitted_ids}

    results = []
    for emp_id, data in emp_data.items():
        if not _is_employee_active(data["records"]):
            continue

        issues = analyse_employee(
            emp_id, data["records"],
            missed_lookup, leave_lookup,
            short_leave_lookup, two_late_lookup,
        )

        total_issues = sum(len(v) for v in issues.values())
        results.append({
            "employee_id":   emp_id,
            "employee_name": data["name"],
            "total_issues":  total_issues,
            "issues": {
                key: [_serialize_record(r) for r in recs]
                for key, recs in issues.items()
            },
        })

    # employees with issues first, then alphabetically
    results.sort(key=lambda x: (-x["total_issues"], (x["employee_name"] or "").lower()))
    return results, from_date, to_date


# ---------------------------------------------------------------------------
# Public whitelisted API
# ---------------------------------------------------------------------------

@frappe.whitelist()
def get_attendance_analysis(from_date, to_date, employees=None):
    """
    Run the 4-check attendance analysis for the given period and return
    JSON-serialisable results.  Called by the Attendance Summary Report page.

    Args:
        from_date: str  "YYYY-MM-DD"
        to_date:   str  "YYYY-MM-DD"
        employees: str  JSON list of employee IDs to restrict analysis to;
                        pass an empty list or omit to analyse all.

    Returns:
        List of dicts, one per active employee, sorted by issue count desc.
    """
    results, _, _ = _run_analysis(from_date, to_date, employees)
    return results


@frappe.whitelist()
def export_attendance_summary_excel(from_date, to_date, employees=None):
    """
    Generate an XLSX workbook for the attendance summary and return it as a
    base64-encoded payload so the browser can trigger a file download.

    Returns:
        {"filename": str, "content_type": str, "content": str (base64)}
    """
    import base64, io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    results, fd, td = _run_analysis(from_date, to_date, employees)

    # ── Style helpers ──────────────────────────────────────────────────────

    def _header_row(ws, hex_color):
        fill = PatternFill("solid", fgColor=hex_color)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left")

    def _col_widths(ws, widths):
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = w

    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append([
        "Employee ID", "Employee Name",
        "Missed Attendance", "Leave Applications",
        "Short Leave", "Two Late \u2192 Half Day",
        "Total Issues",
    ])
    _header_row(ws_sum, "2563EB")
    for emp in results:
        ws_sum.append([
            emp["employee_id"],
            emp["employee_name"],
            len(emp["issues"].get("missed_attendance_request", [])),
            len(emp["issues"].get("leave_application", [])),
            len(emp["issues"].get("short_leave_application", [])),
            len(emp["issues"].get("two_late_to_half_day", [])),
            emp["total_issues"],
        ])
    _col_widths(ws_sum, {
        "A": 16, "B": 30, "C": 20, "D": 22, "E": 14, "F": 24, "G": 14,
    })

    # ── Detail sheets (one per issue type) ────────────────────────────────
    DETAIL_SHEETS = [
        ("missed_attendance_request", "DC2626", "Missed Attendance"),
        ("leave_application",         "EA580C", "Leave Applications"),
        ("short_leave_application",   "2563EB", "Short Leave"),
        ("two_late_to_half_day",      "7C3AED", "Two Late \u2192 Half Day"),
    ]
    DETAIL_HEADERS = [
        "Employee ID", "Employee Name", "Date", "Status",
        "In Time", "Out Time", "Shift", "Remarks", "Document ID",
    ]

    for issue_key, color, sheet_name in DETAIL_SHEETS:
        ws = wb.create_sheet(sheet_name)
        ws.append(DETAIL_HEADERS)
        _header_row(ws, color)
        for emp in results:
            for rec in emp["issues"].get(issue_key, []):
                ws.append([
                    emp["employee_id"],
                    emp["employee_name"],
                    rec.get("attendance_date", ""),
                    rec.get("status") or rec.get("custom_ucsc_status", ""),
                    rec.get("in_time", ""),
                    rec.get("out_time", ""),
                    rec.get("shift", ""),
                    rec.get("custom_remarks", ""),
                    rec.get("name", ""),
                ])
        _col_widths(ws, {
            "A": 16, "B": 30, "C": 14, "D": 20,
            "E": 12, "F": 12, "G": 18, "H": 32, "I": 36,
        })

    # ── Serialise to base64 ────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")

    return {
        "filename":     f"attendance_summary_{fd}_to_{td}.xlsx",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "content":      b64,
    }


@frappe.whitelist()
def export_attendance_summary_pdf(from_date, to_date, employees=None):
    """
    Generate a PDF attendance summary report and return it as a
    base64-encoded payload so the browser can trigger a file download.

    Returns:
        {"filename": str, "content_type": str, "content": str (base64)}
    """
    import base64
    from frappe.utils import now_datetime
    from frappe.utils.pdf import get_pdf

    results, fd, td = _run_analysis(from_date, to_date, employees)
    html = _build_pdf_html(results, fd, td)
    pdf_bytes = get_pdf(html)
    b64 = base64.b64encode(pdf_bytes).decode("ascii")

    return {
        "filename":     f"attendance_summary_{fd}_to_{td}.pdf",
        "content_type": "application/pdf",
        "content":      b64,
    }


def _build_pdf_html(results, from_date, to_date):
    """Build a print-ready HTML string for the PDF attendance summary."""
    from html import escape as esc
    from frappe.utils import now_datetime

    gen_date = str(now_datetime())[:16]
    mdash    = "\u2014"  # used in f-string expressions (backslashes not allowed there)

    # ── Aggregate stats ────────────────────────────────────────────────────
    total_employees = len(results)
    with_issues     = sum(1 for e in results if e["total_issues"] > 0)
    counts = {
        "missed_attendance_request": 0,
        "leave_application":         0,
        "short_leave_application":   0,
        "two_late_to_half_day":      0,
    }
    for emp in results:
        for k in counts:
            counts[k] += len(emp["issues"].get(k, []))

    # ── Summary table rows ─────────────────────────────────────────────────
    summary_rows = ""
    for i, emp in enumerate(results):
        bg   = "#F9FAFB" if i % 2 == 0 else "#FFFFFF"
        tot  = emp["total_issues"]
        clr  = "#DC2626" if tot > 0 else "#166534"
        summary_rows += (
            f'<tr style="background:{bg};">'
            f'<td>{esc(emp["employee_id"])}</td>'
            f'<td>{esc(emp["employee_name"])}</td>'
            f'<td style="text-align:center;">'
            f'{len(emp["issues"].get("missed_attendance_request", []))}</td>'
            f'<td style="text-align:center;">'
            f'{len(emp["issues"].get("leave_application", []))}</td>'
            f'<td style="text-align:center;">'
            f'{len(emp["issues"].get("short_leave_application", []))}</td>'
            f'<td style="text-align:center;">'
            f'{len(emp["issues"].get("two_late_to_half_day", []))}</td>'
            f'<td style="text-align:center;font-weight:700;color:{clr};">{tot}</td>'
            f'</tr>'
        )

    # ── Per-employee detail sections (only employees with issues) ──────────
    ISSUE_META = [
        ("missed_attendance_request", "#DC2626", "Missed Attendance"),
        ("leave_application",         "#EA580C", "Leave Applications"),
        ("short_leave_application",   "#2563EB", "Short Leave"),
        ("two_late_to_half_day",      "#7C3AED", "Two Late \u2192 Half Day"),
    ]

    detail_html = ""
    for emp in results:
        if not emp["total_issues"]:
            continue
        detail_html += (
            f'<div class="emp-block">'
            f'<div class="emp-header">'
            f'{esc(emp["employee_name"])} '
            f'<span style="font-weight:400;color:#6B7280;">({esc(emp["employee_id"])})</span>'
            f'</div>'
        )
        for issue_key, color, label in ISSUE_META:
            recs = emp["issues"].get(issue_key, [])
            if not recs:
                continue
            detail_html += (
                f'<p class="issue-type-label" style="color:{color};">'
                f'{esc(label)} ({len(recs)})</p>'
                f'<table>'
                f'<thead><tr style="background:{color};">'
                f'<th>Date</th><th>Status</th>'
                f'<th>In Time</th><th>Out Time</th>'
                f'<th>Shift</th><th>Remarks</th>'
                f'</tr></thead><tbody>'
            )
            for j, rec in enumerate(recs):
                row_bg   = "#F9FAFB" if j % 2 == 0 else "#FFFFFF"
                st       = esc(rec.get("status") or rec.get("custom_ucsc_status", "\u2014"))
                detail_html += (
                    f'<tr style="background:{row_bg};">'
                    f'<td>{esc(str(rec.get("attendance_date", "") or mdash))}</td>'
                    f'<td>{st}</td>'
                    f'<td>{esc(str(rec.get("in_time",  "") or mdash))}</td>'
                    f'<td>{esc(str(rec.get("out_time", "") or mdash))}</td>'
                    f'<td>{esc(str(rec.get("shift",    "") or mdash))}</td>'
                    f'<td>{esc(str(rec.get("custom_remarks", "") or mdash))}</td>'
                    f'</tr>'
                )
            detail_html += "</tbody></table>"
        detail_html += "</div>"

    # ── Assemble full HTML ─────────────────────────────────────────────────
    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  body {{
    font-family: Arial, sans-serif;
    font-size: 11px;
    color: #1F2937;
    margin: 0;
    padding: 24px 28px;
  }}
  h1 {{
    font-size: 20px;
    color: #1E40AF;
    margin: 0 0 3px;
  }}
  .subtitle {{
    font-size: 11px;
    color: #6B7280;
    margin-bottom: 20px;
  }}
  .stats-grid {{
    display: table;
    width: 100%;
    margin-bottom: 22px;
    border-collapse: separate;
    border-spacing: 8px 0;
  }}
  .stat-cell {{
    display: table-cell;
    width: 16%;
    border: 1px solid #E5E7EB;
    border-radius: 4px;
    text-align: center;
    padding: 8px 4px;
    vertical-align: middle;
  }}
  .stat-val {{
    font-size: 22px;
    font-weight: 700;
    line-height: 1.1;
  }}
  .stat-lbl {{
    font-size: 9px;
    color: #6B7280;
    margin-top: 3px;
    line-height: 1.3;
  }}
  h2 {{
    font-size: 13px;
    color: #1E40AF;
    border-bottom: 2px solid #2563EB;
    padding-bottom: 5px;
    margin: 22px 0 10px;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 14px;
    font-size: 10px;
  }}
  th {{
    color: #FFFFFF;
    padding: 6px 8px;
    text-align: left;
    font-weight: 600;
  }}
  td {{
    padding: 5px 8px;
    border: 1px solid #E5E7EB;
  }}
  .summary-thead th {{
    background: #2563EB;
  }}
  .emp-block {{
    margin-bottom: 18px;
    page-break-inside: avoid;
  }}
  .emp-header {{
    background: #EFF6FF;
    border-left: 4px solid #2563EB;
    padding: 6px 10px;
    font-size: 12px;
    font-weight: 700;
    color: #1E40AF;
    margin-bottom: 6px;
  }}
  .issue-type-label {{
    font-size: 10px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.4px;
    margin: 8px 0 3px;
  }}
  .page-break {{
    page-break-after: always;
  }}
</style>
</head>
<body>

<h1>Attendance Summary Report</h1>
<p class="subtitle">
  Period: <strong>{esc(str(from_date))}</strong> to <strong>{esc(str(to_date))}</strong>
  &nbsp;&nbsp;|&nbsp;&nbsp;Generated: {esc(gen_date)}
</p>

<!-- Statistics -->
<div class="stats-grid">
  <div class="stat-cell" style="border-top:3px solid #2563EB;">
    <div class="stat-val" style="color:#2563EB;">{total_employees}</div>
    <div class="stat-lbl">Employees Analysed</div>
  </div>
  <div class="stat-cell" style="border-top:3px solid #DC2626;">
    <div class="stat-val" style="color:#DC2626;">{with_issues}</div>
    <div class="stat-lbl">With Issues</div>
  </div>
  <div class="stat-cell" style="border-top:3px solid #DC2626;">
    <div class="stat-val" style="color:#DC2626;">{counts["missed_attendance_request"]}</div>
    <div class="stat-lbl">Missed Attendance</div>
  </div>
  <div class="stat-cell" style="border-top:3px solid #EA580C;">
    <div class="stat-val" style="color:#EA580C;">{counts["leave_application"]}</div>
    <div class="stat-lbl">Leave Applications</div>
  </div>
  <div class="stat-cell" style="border-top:3px solid #2563EB;">
    <div class="stat-val" style="color:#2563EB;">{counts["short_leave_application"]}</div>
    <div class="stat-lbl">Short Leave</div>
  </div>
  <div class="stat-cell" style="border-top:3px solid #7C3AED;">
    <div class="stat-val" style="color:#7C3AED;">{counts["two_late_to_half_day"]}</div>
    <div class="stat-lbl">Two Late &rarr; Half Day</div>
  </div>
</div>

<!-- Summary table -->
<h2>Employee Summary</h2>
<table>
  <thead class="summary-thead">
    <tr>
      <th>Employee ID</th>
      <th>Employee Name</th>
      <th style="text-align:center;">Missed</th>
      <th style="text-align:center;">Leave</th>
      <th style="text-align:center;">Short Leave</th>
      <th style="text-align:center;">Two Late</th>
      <th style="text-align:center;">Total</th>
    </tr>
  </thead>
  <tbody>
    {summary_rows}
  </tbody>
</table>

<!-- Detailed breakdown (employees with issues only) -->
<div class="page-break"></div>
<h2>Detailed Breakdown</h2>
{detail_html if detail_html else
 '<p style="color:#6B7280;font-style:italic;">No attendance issues found for this period.</p>'}

</body>
</html>"""


@frappe.whitelist()
def send_attendance_emails(from_date, to_date, employee=None,
                           send_even_if_no_issues=0,
                           selected_employees=None):
    """
    Enqueue a background job to send attendance summary emails for the period.
    Returns immediately; emails are sent asynchronously.
    Restricted to users with the System Manager role.

    Args:
        from_date:               str   "YYYY-MM-DD"
        to_date:                 str   "YYYY-MM-DD"
        employee:                str   optional — send to one employee only
        send_even_if_no_issues:  int   1 to email employees with no issues too
        selected_employees:      str   JSON list of employee IDs to restrict sending to

    Returns:
        {"status": "queued", "message": str}
    """
    import json
    if "System Manager" not in frappe.get_roles():
        frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

    allowed = None
    if selected_employees:
        allowed = json.loads(selected_employees) if isinstance(selected_employees, str) else selected_employees

    frappe.enqueue(
        "attendance_processor.utils.api._do_send_emails",
        from_date=from_date,
        to_date=to_date,
        employee=employee or None,
        send_even_if_no_issues=cint(send_even_if_no_issues),
        selected_employees=allowed,
        queue="long",
        timeout=3600,
    )
    count = len(allowed) if allowed is not None else "all selected"
    return {
        "status":  "queued",
        "message": f"Email job queued for {from_date} to {to_date} "
                   f"({count} employee(s)). "
                   "Emails will be delivered in the background.",
    }


@frappe.whitelist()
def get_email_send_preview(from_date, to_date, employees=None):
    """
    Return the list of employees who would receive an email if Send Emails
    were triggered right now: active employees with at least one issue.
    Restricted to System Manager.

    Returns:
        List of dicts: [{employee_id, employee_name, email, issue_count}, ...]
    """
    import json
    if "System Manager" not in frappe.get_roles():
        frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

    from_date = getdate(from_date)
    to_date   = getdate(to_date)

    employees_list = None
    if employees:
        raw = json.loads(employees) if isinstance(employees, str) else employees
        employees_list = [e for e in raw if e] or None

    missed_lookup      = get_missed_requests_lookup(from_date, to_date)
    leave_lookup       = get_leave_applications_lookup(from_date, to_date)
    short_leave_lookup = get_short_leave_lookup(from_date, to_date)
    two_late_lookup    = get_two_late_lookup(from_date, to_date)
    all_records        = get_attendance_records(from_date, to_date,
                                                employees=employees_list)

    emp_data   = _build_emp_data(all_records)
    active_ids = _get_active_employee_ids()

    # Batch-fetch email (user_id) for all active candidates in one query
    candidate_ids = [eid for eid in emp_data if eid in active_ids]
    email_map = {
        r.name: r.user_id
        for r in frappe.get_all(
            "Employee",
            filters={"name": ["in", candidate_ids]},
            fields=["name", "user_id"],
        )
    } if candidate_ids else {}

    recipients = []
    for emp_id, data in emp_data.items():
        if emp_id not in active_ids:
            continue
        issues = analyse_employee(
            emp_id, data["records"],
            missed_lookup, leave_lookup,
            short_leave_lookup, two_late_lookup,
        )
        total_issues = sum(len(v) for v in issues.values())
        if total_issues == 0:
            continue
        recipients.append({
            "employee_id":   emp_id,
            "employee_name": data["name"],
            "email":         email_map.get(emp_id) or "",
            "issue_count":   total_issues,
        })

    recipients.sort(key=lambda x: (x["employee_name"] or "").lower())
    return recipients


# ---------------------------------------------------------------------------
# Background worker (called via frappe.enqueue — not directly by users)
# ---------------------------------------------------------------------------

def _do_send_emails(from_date, to_date, employee=None,
                    send_even_if_no_issues=False, period_label=None,
                    selected_employees=None):
    """
    Background worker: runs the full attendance analysis and sends summary
    emails.  Errors per employee are logged and do not abort other employees.
    selected_employees: optional list of employee IDs — only these are emailed.
    """
    from_date = getdate(from_date)
    to_date   = getdate(to_date)

    if not period_label:
        period_label = f"{from_date} to {to_date} (Custom)"

    missed_lookup      = get_missed_requests_lookup(from_date, to_date)
    leave_lookup       = get_leave_applications_lookup(from_date, to_date)
    short_leave_lookup = get_short_leave_lookup(from_date, to_date)
    two_late_lookup    = get_two_late_lookup(from_date, to_date)
    all_records        = get_attendance_records(from_date, to_date,
                                                employee=employee)

    emp_data   = _build_emp_data(all_records)
    active_ids = _get_active_employee_ids(employee=employee)
    allowed    = set(selected_employees) if selected_employees else None

    for emp_id, data in emp_data.items():
        if emp_id not in active_ids:
            continue
        if allowed is not None and emp_id not in allowed:
            continue
        try:
            issues = analyse_employee(
                emp_id, data["records"],
                missed_lookup, leave_lookup,
                short_leave_lookup, two_late_lookup,
            )
            send_summary_email(
                emp_id, data["name"], issues, period_label,
                send_even_if_no_issues=bool(send_even_if_no_issues),
            )
        except Exception as exc:
            frappe.log_error(
                f"Error sending email to {emp_id} ({data['name']}): {exc}",
                title="Attendance Summary: Email Send Failed",
            )


# ---------------------------------------------------------------------------
# Manual trigger API — called from the Settings form "Send Now" buttons
# ---------------------------------------------------------------------------

@frappe.whitelist()
def send_test_email_to_employee(employee, period_type="weekly",
                                from_date=None, to_date=None):
    """
    Send a test attendance summary email to a single employee immediately.
    Always sends even if the employee has no issues (for testing purposes).
    Restricted to System Manager.

    Args:
        employee:    str  — Employee document name
        period_type: str  — "weekly", "monthly", or "custom"
        from_date:   str  — required when period_type == "custom" ("YYYY-MM-DD")
        to_date:     str  — required when period_type == "custom" ("YYYY-MM-DD")

    Returns:
        {"status": "sent"|"error", "message": str}
    """
    if "System Manager" not in frappe.get_roles():
        frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

    from frappe.utils import nowdate, add_days, get_first_day

    if period_type == "now":
        today        = getdate(nowdate())
        fd           = today
        td           = today
        period_label = f"{today} (Today)"
    elif period_type == "weekly":
        today          = getdate(nowdate())
        days_since_mon = today.weekday()
        fd             = add_days(today, -(days_since_mon + 7))
        td             = add_days(fd, 6)
        period_label   = f"{fd} to {td} (Weekly)"
    elif period_type == "monthly":
        today      = getdate(nowdate())
        first_this = get_first_day(today)
        last_prev  = add_days(first_this, -1)
        first_prev = get_first_day(last_prev)
        fd           = first_prev
        td           = last_prev
        period_label = last_prev.strftime("%B %Y") + " (Monthly)"
    else:
        if not from_date or not to_date:
            frappe.throw(frappe._("from_date and to_date are required for custom period."))
        fd           = getdate(from_date)
        td           = getdate(to_date)
        period_label = f"{fd} to {td} (Custom)"

    emp = frappe.db.get_value("Employee", employee, ["name", "employee_name"], as_dict=True)
    if not emp:
        frappe.throw(frappe._("Employee {0} not found.").format(employee))

    missed_lookup      = get_missed_requests_lookup(fd, td)
    leave_lookup       = get_leave_applications_lookup(fd, td)
    short_leave_lookup = get_short_leave_lookup(fd, td)
    two_late_lookup    = get_two_late_lookup(fd, td)
    all_records        = get_attendance_records(fd, td, employees=[employee])

    if not all_records:
        # No attendance records but still send to confirm email delivery
        emp_records = []
    else:
        emp_records = [r for r in all_records if r.employee == employee]

    issues = analyse_employee(
        employee, emp_records,
        missed_lookup, leave_lookup,
        short_leave_lookup, two_late_lookup,
    )

    recipient = frappe.db.get_value("Employee", employee, "user_id")
    if not recipient:
        frappe.throw(
            frappe._("Employee {0} has no linked User account. Cannot send email.").format(employee)
        )

    from attendance_processor.utils.email_report import build_html_email
    employee_name = emp.employee_name or employee
    total_issues  = sum(len(v) for v in issues.values())

    if total_issues > 0:
        subject = f"Attendance Summary — {period_label}"
    else:
        subject = f"Attendance Summary — {period_label} (No Issues)"

    html_body = build_html_email(employee_name, issues, period_label)

    try:
        frappe.sendmail(
            recipients=[recipient],
            subject=subject,
            message=html_body,
            now=True,
        )
    except Exception as exc:
        frappe.log_error(
            f"Test email failed for {employee} ({employee_name}): {exc}",
            title="Attendance Summary: Test Email Failed",
        )
        return {"status": "error", "message": str(exc)}

    return {
        "status":  "sent",
        "message": f"Test email sent to {recipient} ({employee_name}) "
                   f"for period {period_label}.",
    }


@frappe.whitelist()
def trigger_weekly_report():
    """
    Enqueue the weekly attendance summary job immediately (on-demand).
    Restricted to System Manager.  Does not affect the last_sent tracking,
    so the scheduled run will still fire on the configured day.

    Returns:
        {"status": "queued", "message": str}
    """
    if "System Manager" not in frappe.get_roles():
        frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

    frappe.enqueue(
        "attendance_processor.scheduler.send_weekly_attendance_summary",
        queue="long",
        timeout=3600,
    )
    return {
        "status":  "queued",
        "message": "Weekly attendance summary job has been queued. "
                   "Emails will be delivered in the background.",
    }


@frappe.whitelist()
def trigger_monthly_report():
    """
    Enqueue the monthly attendance summary job immediately (on-demand).
    Restricted to System Manager.

    Returns:
        {"status": "queued", "message": str}
    """
    if "System Manager" not in frappe.get_roles():
        frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

    frappe.enqueue(
        "attendance_processor.scheduler.send_monthly_attendance_summary",
        queue="long",
        timeout=3600,
    )
    return {
        "status":  "queued",
        "message": "Monthly attendance summary job has been queued. "
                   "Emails will be delivered in the background.",
    }


@frappe.whitelist()
def trigger_approver_summary(from_date=None, to_date=None):
    """
    Enqueue the approver summary job immediately (on-demand).
    When from_date and to_date are provided they are used directly; otherwise
    the Lookback Period from Attendance Processor Settings is applied.
    Restricted to System Manager.

    Args:
        from_date: Optional YYYY-MM-DD string for the period start.
        to_date:   Optional YYYY-MM-DD string for the period end.

    Returns:
        {"status": "queued", "message": str}
    """
    if "System Manager" not in frappe.get_roles():
        frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

    if from_date and to_date:
        fd = getdate(from_date)
        td = getdate(to_date)
        frappe.enqueue(
            "attendance_processor.scheduler.send_approver_attendance_summary",
            queue="long",
            timeout=3600,
            from_date=str(fd),
            to_date=str(td),
        )
        return {
            "status":  "queued",
            "message": f"Approver summary job has been queued for {fd} to {td}. "
                       "Emails will be delivered in the background.",
        }
    else:
        try:
            settings = frappe.get_single("Attendance Processor Settings")
            lookback = int(settings.approver_summary_lookback_days or 90)
        except Exception:
            lookback = 90

        frappe.enqueue(
            "attendance_processor.scheduler.send_approver_attendance_summary",
            queue="long",
            timeout=3600,
            lookback_days=lookback,
        )
        return {
            "status":  "queued",
            "message": f"Approver summary job has been queued (last {lookback} days). "
                       "Emails will be delivered in the background.",
        }


@frappe.whitelist()
def send_test_approver_email(approver_user, from_date=None, to_date=None, lookback_days=None):
    """
    Send a test approver summary email to the specified User immediately.
    Restricted to System Manager.

    Args:
        approver_user:  ERPNext User ID to send the test email to.
        from_date:      Optional explicit from_date (YYYY-MM-DD).
        to_date:        Optional explicit to_date (YYYY-MM-DD).
        lookback_days:  Days to look back when from_date/to_date are omitted.
                        Falls back to the value in Settings (default 90).

    Returns:
        {"status": "sent"|"error", "message": str}
    """
    if "System Manager" not in frappe.get_roles():
        frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

    from frappe.utils import nowdate, add_days
    from attendance_processor.utils.approver_report import (
        fetch_approver_data,
        build_approver_html_email,
        send_approver_summary_email,
    )

    # Resolve date range
    if from_date and to_date:
        fd = getdate(from_date)
        td = getdate(to_date)
        period_label = f"{fd} to {td}"
    else:
        if lookback_days is not None:
            lb = int(lookback_days)
        else:
            try:
                settings = frappe.get_single("Attendance Processor Settings")
                lb = int(settings.approver_summary_lookback_days or 90)
            except Exception:
                lb = 90
        td = getdate(nowdate())
        fd = add_days(td, -lb)
        period_label = f"Last {lb} Days (as of {td})"

    period_label += " (Test)"

    # Resolve approver display name
    approver_name = frappe.db.get_value("User", approver_user, "full_name") or approver_user

    # Fetch data for ALL approvers, then extract only the target approver's slice
    grouped_all = fetch_approver_data(fd, td)
    approver_data = grouped_all.get(approver_user, {
        "approver_name":             approver_name,
        "leave_applications":        [],
        "two_late_applications":     [],
        "short_leave_applications":  [],
        "missed_attendance_requests": [],
    })

    try:
        send_approver_summary_email(
            approver_user,
            approver_name,
            approver_data,
            period_label,
        )
        total = (
            len(approver_data.get("leave_applications", []))
            + len(approver_data.get("two_late_applications", []))
            + len(approver_data.get("short_leave_applications", []))
            + len(approver_data.get("missed_attendance_requests", []))
        )
        return {
            "status":  "sent",
            "message": f"Test approver summary sent to {approver_user} "
                       f"({total} pending item(s) found for the period).",
        }
    except Exception as exc:
        return {
            "status":  "error",
            "message": str(exc),
        }
